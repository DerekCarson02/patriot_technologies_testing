import re
from openpyxl import Workbook
from openpyxl import load_workbook
import sys
import os.path

filePath = r'Z:\Enterasys - Extreme\NPI\4120C\Extreme MAC address'
deletechars = {58: ""}


def loadMacAddresses(macAddressDoc):
    wb = load_workbook(filename=r'{}\{}'.format(filePath, macAddressDoc), data_only=True)
    ws = wb.active
    mac_list = []
    for x in range(0, 770):
        if x % 64 == 0:
            mac_list.append(ws.cell(row=x + 3, column=2).value)

    remove_NoneType = []

    for item in mac_list:
        if item is None:
            continue
        remove_NoneType.append(item)

    out_list = []

    for i in range(len(remove_NoneType)):
        string = ""
        string = remove_NoneType[i]
        out_list.append(string.translate(deletechars))

    return list(out_list)


def createMacDocument(newfileName, mac_Address_List):
    fileCheck = os.path.exists(filePath + newfileName)
    if fileCheck:
        print("{} already exists in this folder! DO NOT overwrite existing work orders!".format(workOrderName))
        sys.exit()

    destination_filename_path = r'{}\{}.xlsx'.format(filePath, newfileName)

    wb = Workbook()

    # Creating worksheet.
    ws1 = wb.active

    for x in range(0, len(mac_Address_List)):
        value = mac_Address_List[x]
        ws1['A{}'.format(x + 1)] = value

    wb.save(filename=destination_filename_path)
    print("File located in " + destination_filename_path + " created.")


createMacDocument("TEST", loadMacAddresses("\MAC Address Increment - Ver2.xlsx"))
