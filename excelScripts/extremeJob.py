from openpyxl import Workbook
from openpyxl import load_workbook
import sys
import os.path

customer = "Extreme Networks"
wDPath = r'C:/Users/dcarson/Documents/Excel Test/'
modelList = ['FGEXT4120C', 'FGEXTE3120', 'FGEXTE3122', 'FGEXTE2122', 'E1120']


def space():
    print("")


def yes_or_no(question):
    reply = str(input(question + ' (y/n): ')).lower().strip()
    if reply[0] == 'y':
        return True
    if reply[0] == 'n':
        return False
    else:
        space()
        print("Please enter (y/n)")
        print(question + ' (y/n): ')
        return False


def workOrderPrompt():
    while True:
        workOrder = input("Please enter your work order number: ")
        if workOrder.isalpha():
            print("This field does not accept letters!")
            pass
        elif len(workOrder) == 0:
            print("Please enter something into the prompt.")
            pass
        elif len(workOrder) > 3:
            print("This field has a character limit of 3!")
            pass
        elif len(workOrder) < 3:
            print("At least 3 characters should be used for this field!")
            pass
        else:
            return "WO00000" + workOrder


def modelNamePrompt():
    while True:
        space()
        for x in range(len(modelList)):
            print("{}).  '{}'".format(x, modelList[x]))
        space()
        try:
            modelSelection = int(input("Please enter your model selection: [0|{}]".format(len(modelList)-1)))
            if modelSelection < 0:
                print("This is not a valid selection!")
                pass
        except ValueError:
            print("This is not a valid selection!")
            pass
        else:
            try:
                if not yes_or_no("Is {} the correct model name?".format(modelList[modelSelection])):
                    pass
                else:
                    return modelList[modelSelection]
            except IndexError:
                print("This is not a valid selection!")
                pass


def createWorkBook(workOrderName, model):
    fileCheck = os.path.exists(wDPath + workOrderName)
    if fileCheck:
        print("{} already exists in this folder! DO NOT overwrite existing work orders!".format(workOrderName))
        sys.exit()

    destination_filename_path = '{}{}.xlsx'.format(wDPath, workOrderName)

    # Creating workbook object.
    wb = Workbook()

    # Creating worksheet.
    ws1 = wb.active
    ws1.title = workOrderName
    ws1['A1'] = workOrderName
    ws1['C1'] = model

    wb.save(filename=destination_filename_path)
    return destination_filename_path


def serialNumberGeneration(model):
    newGenSystems = ['FGEXT4120C', 'FGEXTE3120']
    oldGenSystems = ['FGEXTE3122', 'FGEXTE2122', 'E1120']
    # for modelIndex in newGenSystems:
    #     if newGenSystems[modelIndex] == model:


def model_FGEXT4120C(workbook_destination):
    wb = load_workbook(filename=workbook_destination)
    wb['A3'] = "FG SN#"
    wb['B3'] = "MBSN"
    wb['C3'] = "Admin MAC"
    wb['D3'] = "TANUM"
    wb['E3'] = "Chassis"
    wb['F3'] = "MCX516A-GCAT"
    wb['G3'] = "480 SSD"
    wb['H3'] = "1.92 SSD"
    wb['I3'] = "Extreme MAC"


# createWorkBook(workOrderPrompt(), modelNamePrompt())
